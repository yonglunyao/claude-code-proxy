"""Excel utilities for md2doc skill."""
import math
from io import StringIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.packaging.custom import StringProperty
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import CONFIG
from logger import get_logger
from utils.aigc_mark import get_aigc_signature

LOG = get_logger("ExcelUtils")

config = CONFIG.md2doc_config
# Default settings
column_width_limit = config.get('md2doc_excel_max_column_width_limit', 32)
max_display_char_limit = config.get('md2doc_excel_max_display_char_limit', 160)
unit_row_height = config.get('md2doc_excel_unit_row_height', 15)
adjust_row_height = config.get('md2doc_excel_adjust_row_height', True)
ai_mark_content = config.get('md2doc_ai_mark_content', "内容由AI生成")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def html2df(html_text: str, request_id: str = ""):
    """Convert HTML tables to pandas DataFrames."""
    if '<table>' not in html_text:
        LOG.error(f"request_id=[{request_id}], no valid table in input content")
        raise ValueError("no valid table in input content")
    try:
        return pd.read_html(StringIO(html_text))
    except Exception as e:
        LOG.error(f"request_id=[{request_id}], md2excel failed: {str(e)}")
        raise ValueError(f"md2excel failed: {str(e)}") from e


def df2excel(tables, target_path: str, **kwargs):
    """Write DataFrames to Excel file."""
    request_id = kwargs.get("request_id", "")
    sheet_name_list = kwargs.get('sheet_name_list', [])
    try:
        with pd.ExcelWriter(target_path, engine='xlsxwriter') as writer:
            for i, df in enumerate(tables):
                df.columns = handle_empty_column_header(df.columns)
                if sheet_name_list and len(sheet_name_list) > i:
                    sheet_name = sheet_name_list[i]
                else:
                    sheet_name = f'Sheet{i + 1}'
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    except Exception as e:
        LOG.error(f"request_id=[{request_id}], md2excel failed: {str(e)}")
        raise ValueError(f"md2excel failed: {str(e)}") from e


def handle_empty_column_header(header):
    """Handle empty column headers."""
    return ['' if 'Unnamed: ' in col else col for col in header]


def calculate_display_length(text):
    """Calculate display length (Chinese chars = 2, others = 1)."""
    if text is None:
        return 0
    length = 0
    for char in str(text):
        if '\u4e00' <= char <= '\u9fff' or char in '，。！？、；：""''（）【】《》':
            length += 2
        else:
            length += 1
    return length


def adjust_wb(target_path: str, request_id: str = ""):
    """Adjust workbook formatting."""
    try:
        wb = load_workbook(target_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            adjust_ws(ws)
        wb.save(target_path)
    except Exception as e:
        LOG.warning(f"request_id=[{request_id}], adjust_wb failed: {str(e)}")


def set_excel_properties(target_path: str, content: str, request_id: str = ""):
    """Set Excel AIGC properties."""
    try:
        wb = load_workbook(target_path)
        custom_props = wb.custom_doc_props
        if custom_props is not None:
            aigc_signature = get_aigc_signature(content, request_id)
            custom_props.append(StringProperty(name='AIGC', value=aigc_signature))
        wb.save(target_path)
    except Exception as e:
        LOG.error(f'request_id={request_id}, Update excel metadata error: {str(e)}')


def adjust_ws(ws):
    """Adjust worksheet formatting."""
    adjust_column(ws)
    if adjust_row_height:
        adjust_row(ws)
    auto_wrap_text(ws)

def auto_wrap_text(ws):
    """Enable text wrapping and set borders."""
    alignment = Alignment(wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            set_cell_border(cell)
            cell.alignment = alignment
            set_formula_to_str(cell)


def set_formula_to_str(cell):
    """Convert non-formula cells with 'f' type to string."""
    value = cell.value
    if value is not None and cell.data_type == 'f' and not value.startswith('='):
        cell.data_type = 's'


def set_cell_border(cell):
    """Set thin border for cell."""
    cell.border = thin_border


def adjust_row(ws):
    """Adjust row heights."""
    for row in ws.iter_rows():
        max_width = get_row_max_width(row)
        ws.row_dimensions[row[0].row].height = math.ceil(max_width / column_width_limit) * unit_row_height


def get_row_max_width(row):
    """Get maximum display width in a row."""
    cell_str_length = 0
    for cell in row:
        cell_length = calculate_display_length(cell.value)
        if cell_length > cell_str_length:
            cell_str_length = cell_length
        if cell_str_length >= max_display_char_limit:
            cell_str_length = max_display_char_limit
            break
    return cell_str_length


def adjust_column(ws):
    """Adjust column widths."""
    for column in ws.columns:
        max_width = get_column_max_width(column)
        adjusted_width = (max_width + 2)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width


def get_column_max_width(column):
    """Get maximum display width in a column."""
    max_width = 0
    for cell in column:
        try:
            cell_length = calculate_display_length(cell.value)
            if cell_length > max_width:
                max_width = cell_length
            if max_width >= column_width_limit:
                max_width = column_width_limit
                break
        except Exception as e:
            LOG.warning(f"get_column_max_width failed: {str(e)}")
            pass
    return max_width
